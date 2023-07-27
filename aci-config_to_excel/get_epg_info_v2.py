import requests
import urllib3
import openpyxl

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

host = "192.168.1.119"
username = "admin"
password = "cisco123"

def authentication():
    print("\nAuthenticating with the device . . .\n")
    # Get an authentication token for the device
    url = "https://" + host + "/api/aaaLogin.json"
    # Login data
    data = """
    {
        "aaaUser": {
                "attributes": {
                "name": \"""" + username + """\",
                "pwd": \"""" + password + """\"
            }
        }
    }
    """
    #print(data)

    response = requests.post(url, data=data, headers={'Content-Type': 'application/json'},verify=False)

    # print json.dumps(response.json(), indent=2)
    if response.status_code == requests.codes.ok:
        print("Authentication successful!\n")
    else:
        print("Authentication failed! Please verify login credentials!\n")
        exit(0)

    token = response.json()['imdata'][0]['aaaLogin']['attributes']['token']
    #print("Authentication Token: " + token + "\n")

    return token


def get_tenant(token):
    url = "https://" + host + "/api/class/fvTenant.json"
    cookie = {'APIC-Cookie': token}
    response = requests.get(url, cookies=cookie, headers={'Content-Type':'application/json'}, verify=False)

    if response.status_code == requests.codes.ok:
        # Get Tenant Name
        totalcount = int(response.json()['totalCount'])
        for i in range(totalcount):
            tenantname = response.json()['imdata'][i]['fvTenant']['attributes']['name']
            print(tenantname)
    else:
        print("Failed!\n")
        exit(0)


def get_epg(token):
    print("\nGetting EPG info . . .\n")
    url = "https://" + host + '/api/node/mo/uni/tn-Tenant_ATX/ap-AP_ATX.json?query-target=subtree&target-subtree-class=fvAEPg'

    cookie = {'APIC-Cookie': token}
    response = requests.get(url, cookies=cookie, headers={'Content-Type':'application/json'}, verify=False)

    epg_list = []

    if response.status_code == requests.codes.ok:
        dict_epg = response.json()['imdata']
        #print(dict_epg)
        for epg in dict_epg:
            epg_list.append(epg['fvAEPg']['attributes']['dn'])
            #print(epg['fvAEPg']['attributes']['dn'])
    else:
        print("Failed!\n")
        exit(0)

#    print(epg_list)
    return epg_list



def get_node_port(token, epg_list):
    print("\nGetting Node info . . .\n")

    dict_epg_node = {}
    list_epg_node = []

    for epg in epg_list:
        
        url = "https://" + host + "/api/node/mo/uni/epp/fv-[" + epg + "].json?query-target=subtree&target-subtree-class=fvIfConn"

        cookie = {'APIC-Cookie': token}
        response = requests.get(url, cookies=cookie, headers={'Content-Type':'application/json'}, verify=False)

        if response.status_code == requests.codes.ok:
            dict_node = response.json()['imdata']
            #print(dict_node)
            for node in dict_node:
                #print (node['fvIfConn']['attributes']['dn'])
                epg_port = node['fvIfConn']['attributes']['dn']
                list_epg_node.append(epg_port)
        else:
            print("Failed!\n")
            exit(0)
        
    #print(dict_epg_node)
    #print(list_epg_node)
#    return dict_epg_node
    return list_epg_node

def excel_export(list_epg_node):
    print("\nExport Data to Excel . . .\n")

    # Open Excel
    wb = openpyxl.load_workbook('format.xlsx')
    ws = wb['Sheet1']
    i = 2

    for k in list_epg_node:
        vpc = ''
        port = ''

        en_start = k.rfind('epg-')
        en_end = k.rfind(']/node')
        epg = k[en_start+4:en_end]
        #print(epg)

        nn_start = k.find('node-')
        nn_end = k.find('/stpathatt')
        node = k[nn_start+5:nn_end]
        #print(node)

        if k.find('/stpathatt-[eth') != -1 :
            pn_start = k.find('eth')
            pn_end = k.find(']/conndef')
            port = k[pn_start+3:pn_end]
            #print(port)

        else:
            vpcn_start = k.find('stpathatt-[')
            vpcn_end = k.find(']/conndef')
            vpc = k[vpcn_start+11:vpcn_end]
            #print(vpc)

        vn_start = k.find('vlan')
        vn_end = k.find(']-[')
        vlan = k[vn_start+5:vn_end]
        #print(vlan)

        # Write Data in Excel   
        ws.cell(i,1,value = epg)  
        ws.cell(i,2,value = node)
        ws.cell(i,3,value = port)
        ws.cell(i,4,value = vpc)
        ws.cell(i,5,value = vlan)

        i = i + 1

    wb.save('format.xlsx')  



if __name__ == "__main__":
    token = authentication()
    #get_tenant(token)
    epg_list = get_epg(token)
    list_epg_node = get_node_port(token, epg_list)
    excel_export(list_epg_node)

