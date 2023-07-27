# import
import openpyxl
import yaml

########################################################
# Obtain Excel sheet name
########################################################
def excel_sheet(filedir,filename):

    xlsx_file = filedir + r'/' + filename
    yaml_file = filedir + r'/' + 'main.yml'
    wb = openpyxl.load_workbook(xlsx_file)
    ws_names = wb.sheetnames

    return ws_names


########################################################
# bind definition from xlsx to yaml
########################################################
def bind_yaml(filedir,filename):
    
    xlsx_file = filedir + r'/' + filename
    yaml_file = filedir + r'/' + 'main.yml'
    wb = openpyxl.load_workbook(xlsx_file)
    ws_names = wb.sheetnames
    sheet = wb['static_bind']
    sheet3 = wb['static_bind_vpc']
    sheet2 = wb['epgs']
    sheet4 = wb['accessport']
    yaml_list = []
    pod_id = 1
    epgs_dic = {}

    # epg - encap辞書オブジェクトの作成
    for row2 in sheet2.iter_rows(min_row=2):
        epgs_list = []
        
        for num in range(2):
            epgs_list.append(row2[num].value)

        epgs_dic[epgs_list[0]]= epgs_list[1]

    # epgからencapの解決    
    for row in sheet.iter_rows(min_row=2,min_col=3,max_col=3):       
        sheet.cell((row[0].row),6).value=epgs_dic.get(row[0].value)
        sheet.cell((row[0].row),7).value=pod_id


    # epgからencapの解決2
    for row in sheet3.iter_rows(min_row=2,min_col=3,max_col=3):       
        sheet3.cell((row[0].row),6).value=epgs_dic.get(row[0].value)
        sheet3.cell((row[0].row),7).value=pod_id

    wb.save(xlsx_file)        

    return


########################################################
# default parameter set to yaml
########################################################
def default_set(filedir):
    yaml_file = filedir + r'/' + 'main.yml'
    yaml_list = {
            "username":"admin",
            "password":"cisco123",
            "tenant":"Tenant_ATX",
            "vrf":"VRF_ATX",
            "ap":"AP_ATX",
        }

    with open(yaml_file,'a',) as file:
            yaml.dump(yaml_list, file, default_flow_style=False)
    return

########################################################
# epg definition from xlsx to yaml
########################################################
def epg_yaml(filedir,filename,sheet_name):
    
    xlsx_file = filedir + r'/' + filename
    yaml_file = filedir + r'/' + 'main.yml'
    wb = openpyxl.load_workbook(xlsx_file)
    sheet = wb[sheet_name]
    yaml_list = []

    for row in sheet.rows:
        param_list = []
        yml = {}
        num = 0

        for cell in row:
            param_list.append(cell.value)

        if cell.row == 1:
            key = param_list

        else:
            for i in key:
                yml[i] = param_list[num]
                num += 1

            yaml_list.append(yml)

        yaml_list2 = {
            sheet_name:yaml_list
        }
            
    with open(yaml_file,'a',) as file:
            yaml.dump(yaml_list2, file, default_flow_style=False)
    return

########################################################
# contract definition from xlsx to yaml
########################################################
def contract_yaml(filedir,filename,sheet_name):
    
    xlsx_file = filedir + r'/' + filename
    yaml_file = filedir + r'/' + 'main.yml'
    wb = openpyxl.load_workbook(xlsx_file)
    ws_names = wb.sheetnames
    sheet = wb[sheet_name]

    yaml_list = []
    num_cont = 2

    for row in sheet.rows:
        param_list = []
        yml = {}
        num = 0
        
        for num in range(5):
            param_list.append(row[num].value)

        if row[num].row != 1:
            if param_list[0] == 'vzAny' and param_list[1] == 'vzAny':
                cont_name = 'CONT_vzAny_to_vzAny'

            elif param_list[0] == 'vzAny':
                cont_name = 'CONT_from_vzAny'

            elif param_list[1] == 'vzAny':
                cont_name = 'CONT_to_vzAny'
                
            else:
                cont_name = 'CONT_' + param_list[0][4:]  +'_' + param_list[1][4:]

            row[5].value = cont_name

            sheet2 = wb['contracts']
            sheet2.cell(num_cont,1).value = cont_name
            sheet2.cell(num_cont,2).value = 'subj'
            sheet2.cell(num_cont,3).value = param_list[2]

            sheet3 = wb['filters']
            sheet3.cell(num_cont,1).value = param_list[2]
            sheet3.cell(num_cont,2).value = param_list[2]
            sheet3.cell(num_cont,3).value = param_list[3]
            sheet3.cell(num_cont,4).value = param_list[4]
            num_cont += 1

    wb.save(xlsx_file)

    return

########################################################
# epg-contract definition from xlsx to yaml
########################################################
def epg_cont_yaml(filedir,filename):
    
    xlsx_file = filedir + r'/' + filename
    yaml_file = filedir + r'/' + 'main.yml'
    wb = openpyxl.load_workbook(xlsx_file)
    ws_names = wb.sheetnames
    sheet = wb['flow']
    ids = ['epg','contract_type','contract']
    cont_type = ['consumer','provider']
    yaml_list = []
    
    for row in sheet.iter_rows(min_row=2):
        param_list = []
        yml = {}
        yml2 = {}
        
        for num in [0,1,5]:
            param_list.append(row[num].value)

        #dictionary ymlにEPG、Conacパラメーターを設定        
        yml[ids[0]] = param_list[0]
        yml[ids[1]] = cont_type[0]
        yml[ids[2]] = param_list[2]

        yml2[ids[0]] = param_list[1]
        yml2[ids[1]] = cont_type[1]
        yml2[ids[2]] = param_list[2]
        
        yaml_list.append(yml)
        yaml_list.append(yml2)
        
        yaml_list2 = {
            'epg_contracts':yaml_list
        }
            
    with open(yaml_file,'a',) as file:
           yaml.dump(yaml_list2, file, default_flow_style=False)
    return


########################################################
# main program
########################################################
def main():
    filedir = r'C:\Automation'
    filename = 'Parameter.xlsx'

    # Make excel sheet list 
    sheets = excel_sheet(filedir,filename)

    # EPG encap info distribute to binding
    bind_yaml(filedir,filename)

    # Make yaml file
    default_set(filedir)
    for sheet_name in sheets:
        if sheet_name != 'flow':
            epg_yaml(filedir , filename, sheet_name)
        else:
            contract_yaml(filedir,filename,sheet_name)
    epg_cont_yaml(filedir,filename)


if __name__ == '__main__':
    main()
